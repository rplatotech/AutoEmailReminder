#!/usr/bin/python
"""
The code does the following:

Imports the necessary libraries: win32com.client.
Defines a list of email addresses that should receive the reminder.
The code creates an instance of the Microsoft Outlook application using win32.Dispatch('outlook.application').
For each email address in list_of_emails, the code creates a new email item, sets the recipient and subject,
and sets the body of the email to a reminder message. The email is then sent using mail.Send().
After sending the emails, the code run is completed.

"""

# external library
import win32com.client as win32
# Library for reading data from excel
import openpyxl

# Load the workbook object by providing the path to the excelsheet
wb = openpyxl.load_workbook('test_data.xlsx')

# Select the active sheet
sheet = wb.active
# Get the maximum row number
max_rows = sheet.max_row
# Declare empty list of emails
list_of_emails = []

# Loop through each row and add the value in list
for i in range(2, max_rows + 1):
    cell_obj = sheet.cell(row=i, column=1)
    list_of_emails.append(cell_obj.value)

# # add to this list, all the email IDs that needs to be reminded to fill timesheet
# list_of_emails = ['a@gmail.com', 'b@yahoo.com', ]


def send_email_reminder():
    outlook = win32.Dispatch('outlook.application')
    with open("sent_emails.txt", "a") as f:
        for email in list_of_emails:
            mail = outlook.CreateItem(0)
            mail.To = email
            print(f'Reminder sent to email: {email}', file=f)
            mail.Subject = 'Reminder: Timesheet Submission'
            mail.Body = """Hi team, \n\nAs today is the last day of the week, please ensure to submit your timesheet by today! \n\nHave a great weekend!
            """
            mail.Send()


if __name__ == '__main__':
    send_email_reminder()